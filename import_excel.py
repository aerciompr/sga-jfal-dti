import openpyxl
import json
import sys
import re

def clean_cpf(cpf_str):
    if not cpf_str:
        return ""
    cpf_str = str(cpf_str).strip()
    # If it contains a slash, it's a CNPJ, not a CPF
    if "/" in cpf_str:
        return ""
    # Remove all non-digits to check length
    digits = re.sub(r'\D', '', cpf_str)
    if len(digits) == 11:
        # Format as xxx.xxx.xxx-xx
        return f"{digits[0:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:11]}"
    return cpf_str if len(cpf_str) > 0 and cpf_str != "-" else ""

def parse_excel(file_path):
    try:
        with open(file_path, "rb") as f:
            wb = openpyxl.load_workbook(f, read_only=True)
            sheet = wb.active
            
            # We will group by Process number to merge CNPJ and CPF rows
            processes = {}
            
            # Read header row
            headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
            
            # Find column indices
            idx_processo = -1
            idx_documento = -1
            idx_perito = -1
            idx_cpf_perito = -1
            idx_data = -1
            idx_nome_periciado = -1
            
            for idx, h in enumerate(headers):
                if not h:
                    continue
                h_lower = h.lower()
                if "processo" in h_lower:
                    idx_processo = idx
                elif ("documento" in h_lower or "parte" in h_lower) and "tipo" not in h_lower and "perito" not in h_lower:
                    if "documento" in h_lower:
                        idx_documento = idx
                    elif idx_documento == -1:
                        idx_documento = idx
                elif "perito" in h_lower and "nome" in h_lower:
                    idx_perito = idx
                elif "perito" in h_lower and "cpf" in h_lower:
                    idx_cpf_perito = idx
                elif "perícia" in h_lower and "data" in h_lower and "fim" not in h_lower and "última" not in h_lower and "ultima" not in h_lower:
                    idx_data = idx
                
                # Check for patient name column
                if "nome" in h_lower and "perito" not in h_lower and "especialidade" not in h_lower:
                    idx_nome_periciado = idx
            
            # Fallback to defaults if headers match index
            if idx_processo == -1: idx_processo = 4
            if idx_documento == -1: idx_documento = 5
            if idx_perito == -1: idx_perito = 11
            if idx_data == -1: idx_data = 7
            
            row_idx = 0
            consecutive_empty = 0
            for row in sheet.iter_rows(values_only=True):
                if row_idx == 0:
                    row_idx += 1
                    continue
                
                processo = row[idx_processo]
                if not processo:
                    consecutive_empty += 1
                    if consecutive_empty > 50:
                        break
                    continue
                
                consecutive_empty = 0
                
                documento = row[idx_documento]
                perito = row[idx_perito]
                cpf_perito_val = row[idx_cpf_perito] if (idx_cpf_perito != -1 and len(row) > idx_cpf_perito) else ""
                data_val = row[idx_data]
                
                processo = str(processo).strip()
                perito = str(perito).strip() if perito else ""
                cpf_perito = clean_cpf(cpf_perito_val)
                
                # Check if document is a CNPJ (contains /)
                documento_str = str(documento).strip() if documento else ""
                is_cnpj = "/" in documento_str
                
                if is_cnpj:
                    cpf = ""
                    nome_periciado = ""
                else:
                    cpf = clean_cpf(documento_str)
                    nome_periciado = str(row[idx_nome_periciado]).strip() if (idx_nome_periciado != -1 and row[idx_nome_periciado]) else ""
                
                data_str = ""
                if data_val:
                    if hasattr(data_val, "strftime"):
                        data_str = data_val.strftime("%Y-%m-%d")
                    else:
                        val_str = str(data_val).strip().split(" ")[0]
                        match_ymd = re.match(r'^(\d{4})[-/](\d{2})[-/](\d{2})', val_str)
                        if match_ymd:
                            data_str = f"{match_ymd.group(1)}-{match_ymd.group(2)}-{match_ymd.group(3)}"
                        else:
                            match_dmy = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})', val_str)
                            if match_dmy:
                                data_str = f"{int(match_dmy.group(3)):04d}-{int(match_dmy.group(2)):02d}-{int(match_dmy.group(1)):02d}"
                            else:
                                data_str = val_str
                
                if processo not in processes:
                    processes[processo] = {
                        "processo": processo,
                        "perito": perito,
                        "cpf_perito": cpf_perito,
                        "cpf": cpf,
                        "periciado": nome_periciado,
                        "data_pericia": data_str
                    }
                else:
                    # Update fields if not already populated
                    if cpf and not processes[processo]["cpf"]:
                        processes[processo]["cpf"] = cpf
                    if nome_periciado and not processes[processo]["periciado"]:
                        processes[processo]["periciado"] = nome_periciado
                    if perito and not processes[processo]["perito"]:
                        processes[processo]["perito"] = perito
                    if cpf_perito and not processes[processo]["cpf_perito"]:
                        processes[processo]["cpf_perito"] = cpf_perito
                    if data_str and not processes[processo]["data_pericia"]:
                        processes[processo]["data_pericia"] = data_str
                        
                row_idx += 1
                
            wb.close()
            import gc
            gc.collect()
            return list(processes.values())
            
    except Exception as e:
        return [{"error": str(e)}]

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps([]))
        sys.exit(0)
        
    file_path = sys.argv[1]
    res = parse_excel(file_path)
    print(json.dumps(res, indent=4, ensure_ascii=True))
